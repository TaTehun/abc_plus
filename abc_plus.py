import openpyxl

FILE_NAME = "abc_plus.xlsx"

wb = openpyxl.load_workbook(FILE_NAME)
ws = wb.active

ws["C1"] = "c value"

row = 2
while ws[f"A{row}"].value is not None:
    a = ws[f"A{row}"].value
    b = ws[f"B{row}"].value
    ws[f"C{row}"] = a + b
    row += 1

wb.save(FILE_NAME)
print("완료")
